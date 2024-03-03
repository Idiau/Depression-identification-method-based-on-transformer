from flask import Flask, render_template, request, jsonify, make_response
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from scipy.special import softmax
from flask_ngrok import run_with_ngrok
from flask import Flask
from flask_cors import CORS
import os
import uuid
from urllib.parse import urljoin
import openpyxl
import numpy as np
import json
import pandas as pd

app = Flask(__name__, static_url_path='/static')
# run_with_ngrok(app)  # starts ngrok when the app is run
CORS(app, resources=r'/*')
ALLOWED_EXTENSIONS = {'txt', 'csv', 'xlsx'}
UPLOAD_FOLDER = 'uploads'


def app_test(sentence):
    model_path = "depression-bert-base-cased"
    model = AutoModelForSequenceClassification.from_pretrained(model_path, num_labels=2)
    tokenizer = AutoTokenizer.from_pretrained(model_path)
    sample_tokens = tokenizer(sentence, return_tensors="pt", padding="max_length", truncation=True, max_length=512)
    sample_out = model(**sample_tokens)
    scores = sample_out[0][0].detach().numpy()
    scores = softmax(scores)
    model_prediction = [float(scores[0]), float(scores[1])]
    print("depressive score:", scores[1], "neutral score:", scores[0])
    return model_prediction


# 检查后缀名是否为允许的文件
def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# 获取文件名
def random_filename(filename):
    ext = os.path.splitext(filename)[-1]
    return uuid.uuid4().hex + ext


def get_max_row(sheet):
    i = sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i - 1
        else:
            real_max_row = i
            break

    return real_max_row


@app.route('/')
def home():
    return render_template("index.html")
    # return "<h1>Running Flask!</h1>"


@app.route('/api/test')
def test():
    result_text = {"statusCode": 200, "message": "Success"}
    response = make_response(jsonify(result_text))
    return response


@app.route('/api/upload', methods=['GET', 'POST', 'PUT'])
def upload():
    file = request.files.get("file")
    if file and allowed_file(file.filename):
        filename = random_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(os.path.join(app.root_path, filepath))
        file_url = urljoin(request.host_url, filepath)
        result_text = {"statusCode": 200, "message": filename}
        response = make_response(jsonify(result_text))
    else:
        result_text = {"statusCode": 400, "message": "Error"}
        response = make_response(jsonify(result_text))
    return response


@app.route('/api/detect', methods=['GET', 'POST'])  # decorator
def detect():
    print(request.form.get("file_name"))
    print(request.form.get("text"))
    model_prediction = 2
    filename = request.form.get("file_name")
    text = request.form.get("text")
    path = "./uploads/" + filename
    data = ""
    if text != "":
        data = text
    elif filename.split('.')[1] == 'txt':
        f = open(path, "r+", encoding='utf-8')
        filecontent = f.read()
        data = str(filecontent)
    elif filename.split('.')[1] == 'xlsx':
        data_excel = openpyxl.load_workbook(path)  # 这里可以直接读文件对象
        data_sheet = data_excel[data_excel.sheetnames[0]]  # TODO 这是是读取第一个sheet的数据
        maxRows = get_max_row(data_sheet)  # 行数
        for i in range(1, maxRows + 1):
            data += data_sheet.cell(i, 1).value
    elif filename.split('.')[1] == 'csv':
        data_csv = pd.read_csv(path)
        col_1 = data_csv.iloc[:, 0]  # 获取一列，用一维数据
        str_arr = np.array_str(np.array(col_1))
        data = str_arr.replace('[', '').replace(']', '').replace('\n', '')
    else:
        result_text = {"statusCode": 400, "message": "Error"}
        response = make_response(jsonify(result_text))
        return response
    model_prediction = app_test(data)
    # json_data = json.dumps(float(model_prediction))
    result_text = {"statusCode": 200, "message": model_prediction}
    response = make_response(jsonify(result_text))
    return response


if __name__ == "__main__":
    app.run()
